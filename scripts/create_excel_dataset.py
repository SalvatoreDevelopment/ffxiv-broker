#!/usr/bin/env python3
"""
Script per creare dataset Excel (.xlsx) direttamente
"""

import asyncio
import json
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configurazione
UNIVERSALIS_BASE = "https://universalis.app/api"
XIVAPI_BASE = "https://xivapi.com"
import os
WORLD = os.environ.get("FFXIV_WORLD", "Phoenix")
OUTPUT_FILE = "data/ffxiv_market_dataset.xlsx"
# TEST_ITEMS sara popolato dinamicamente dalla lista tradeable


async def get_item_name(item_id: int) -> str:
    """Ottiene il nome dell'item da XIVAPI"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{XIVAPI_BASE}/item/{item_id}")
            response.raise_for_status()
            data = response.json()
            return data.get("Name", f"Item_{item_id}")
    except Exception:
        return f"Item_{item_id}"


async def get_market_data(item_id: int, world: str) -> Dict[str, Any]:
    """Ottiene i dati di mercato per un item specifico"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{UNIVERSALIS_BASE}/v2/{world}/{item_id}")
            response.raise_for_status()
            data = response.json()
            
            listings = data.get("listings", [])
            history = data.get("recentHistory", [])
            
            # Calcola metriche
            lowest_price = min((listing["pricePerUnit"] for listing in listings), default=None)
            
            # Prezzo medio degli ultimi 7 giorni (filtra per timestamp)
            seven_days_ago = int((datetime.now(timezone.utc) - timedelta(days=7)).timestamp())
            recent_sales = [
                sale for sale in history if int(sale.get("timestamp", 0)) >= seven_days_ago
            ]
            avg_price = (
                sum(sale["pricePerUnit"] for sale in recent_sales) / len(recent_sales)
                if recent_sales
                else None
            )
            
            # Vendite giornaliere (ultimi 7 giorni)
            sales_per_day = len(recent_sales) / 7.0 if recent_sales else 0.0
            
            # Calcola %diff (variazione percentuale rispetto alla media)
            price_diff_percent = None
            if lowest_price is not None and avg_price is not None and avg_price > 0:
                price_diff_percent = round(
                    ((lowest_price - avg_price) / avg_price) * 100, 2
                )
            
            # Flags
            flags = []
            if len(listings) > 10 and sales_per_day < 2:
                flags.append("saturo")
            if (
                lowest_price is not None
                and avg_price is not None
                and lowest_price < avg_price * 0.8
            ):
                flags.append("flip")
            if price_diff_percent is not None and price_diff_percent < -20:
                flags.append("sottovalutato")
            if price_diff_percent is not None and price_diff_percent > 20:
                flags.append("sopravvalutato")
            
            return {
                "item_id": item_id,
                "lowest_price": lowest_price,
                "avg_price_7d": round(avg_price, 2) if avg_price is not None else None,
                "price_diff_percent": price_diff_percent,
                "sales_per_day_7d": round(sales_per_day, 2),
                "flags": ",".join(flags) if flags else "",
                "listings_count": len(listings),
                "history_count": len(history)
            }
    except Exception as e:
        print(f"❌ Errore per item {item_id}: {e}")
        return {
            "item_id": item_id,
            "lowest_price": None,
            "avg_price_7d": None,
            "sales_per_day_7d": 0.0,
            "flags": "error",
            "listings_count": 0,
            "history_count": 0
        }


def create_excel_file(results: List[Dict[str, Any]], filename: str):
    """Crea file Excel con formattazione"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FFXIV Market Data"
    
    # Intestazioni
    headers = [
        "ID Item", "Nome Item", "Prezzo Piu Basso", "Prezzo Medio 7g", "Diff Percentuale",
        "Vendite Giornaliere", "Flags", "Numero Annunci", "Numero Vendite"
    ]
    

    # Stile intestazioni
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Scrivi intestazioni
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Scrivi dati
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result["item_id"])
        ws.cell(row=row, column=2, value=result["item_name"])
        ws.cell(row=row, column=3, value=result["lowest_price"] if result["lowest_price"] is not None else None)
        ws.cell(row=row, column=4, value=result["avg_price_7d"] if result["avg_price_7d"] is not None else None)
        ws.cell(row=row, column=5, value=result["price_diff_percent"] if result["price_diff_percent"] is not None else None)
        ws.cell(row=row, column=6, value=result["sales_per_day_7d"])
        ws.cell(row=row, column=7, value=result["flags"])
        ws.cell(row=row, column=8, value=result["listings_count"])
        ws.cell(row=row, column=9, value=result["history_count"])
        
        # Colora righe con opportunita
        if "flip" in result["flags"]:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FF90EE90", end_color="FF90EE90", fill_type="solid"
                )
        elif "saturo" in result["flags"]:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="FFFFB6C1", end_color="FFFFB6C1", fill_type="solid"
                )
    
    # Auto-adjust column widths
    for col in range(1, 10):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(results) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Salva file
    wb.save(filename)
    print(f"✅ File Excel salvato: {filename}")


async def get_marketable_items() -> List[int]:
    """Ottiene la lista degli item tradeable da Universalis"""
    async with httpx.AsyncClient() as client:
        response = await client.get(f"{UNIVERSALIS_BASE}/marketable")
        response.raise_for_status()
        return response.json()


async def main():
    """Funzione principale"""
    print("📊 Creazione dataset Excel FFXIV Market...")
    
    # Ottieni lista item tradeable
    print("📋 Ottenendo lista item tradeable...")
    marketable_items = await get_marketable_items()
    print(f"✅ Trovati {len(marketable_items)} item tradeable")
    
    # Seleziona un subset per il test
    test_items = marketable_items[:8]  # Primi 8 item tradeable
    print(f"📋 Testando {len(test_items)} item: {test_items}")
    
    results = []
    
    for item_id in test_items:
        print(f"🔄 Processando item {item_id}...")
        
        # Ottieni dati di mercato
        market_data = await get_market_data(item_id, WORLD)
        
        # Ottieni nome item
        item_name = await get_item_name(item_id)
        market_data["item_name"] = item_name
        
        results.append(market_data)
        
        # Pausa tra richieste
        await asyncio.sleep(0.5)
        
        print(f"✅ Item {item_id} ({item_name}): {market_data['lowest_price']} gil, flags: {market_data['flags']}")
    
    # Crea file Excel
    create_excel_file(results, OUTPUT_FILE)
    
    print(f"\n✅ Test completato! File Excel salvato: {OUTPUT_FILE}")
    
    # Statistiche
    with_data = [r for r in results if r["lowest_price"] is not None]
    print(f"📊 Item con dati: {len(with_data)}/{len(results)}")
    print(f"💰 Item con flip: {len([r for r in with_data if 'flip' in r['flags']])}")
    print(f"⚠️ Item saturi: {len([r for r in with_data if 'saturo' in r['flags']])}")


if __name__ == "__main__":
    asyncio.run(main())
