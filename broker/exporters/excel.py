from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _auto_fit(ws: Worksheet) -> None:
    widths: dict[int, int] = {}
    for row in ws.rows:
        for cell in row:
            value = str(cell.value) if cell.value is not None else ""
            # Some cells can be MergedCell without col_idx; fallback to column index
            col_idx = int(getattr(cell, "col_idx", getattr(cell, "column", 0)))
            widths[col_idx] = max(widths.get(col_idx, 0), len(value) + 2)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(60, width)


def build_workbook(
    items: Iterable[dict[str, Any]],
    recipes: Iterable[dict[str, Any]],
    advice: Iterable[dict[str, Any]],
    world: str,
) -> Workbook:
    wb = Workbook()
    ws_items = cast(Worksheet, wb.active)
    ws_items.title = "Items"
    ws_recipes = cast(Worksheet, wb.create_sheet("Recipes"))
    ws_advice = cast(Worksheet, wb.create_sheet("Advisor"))

    # Items sheet
    items_headers = ["item_id", "world", "lowest", "avg_price_7d", "sales_per_day_7d", "flags"]
    ws_items.append(items_headers)
    for h in ws_items[1]:
        h.font = Font(bold=True)
    for it in items:
        ws_items.append(
            [
                it.get("item_id"),
                it.get("world"),
                it.get("lowest"),
                it.get("avg_price_7d"),
                it.get("sales_per_day_7d"),
                ",".join(it.get("flags", [])),
            ]
        )
    ws_items.auto_filter.ref = ws_items.dimensions
    ws_items.freeze_panes = "A2"
    _auto_fit(ws_items)

    # Recipes sheet
    ws_recipes.append(["result_item_id", "amount_result", "ingredients_json"])
    for h in ws_recipes[1]:
        h.font = Font(bold=True)
    for r in recipes:
        ws_recipes.append(
            [
                r.get("result_item_id"),
                r.get("amount_result"),
                r.get("ingredients"),
            ]
        )
    ws_recipes.auto_filter.ref = ws_recipes.dimensions
    ws_recipes.freeze_panes = "A2"
    _auto_fit(ws_recipes)

    # Advice sheet
    ws_advice.append(["item_id", "name", "roi", "sales_per_day", "score", "flags", "risk"])
    for h in ws_advice[1]:
        h.font = Font(bold=True)
    for a in advice:
        ws_advice.append(
            [
                a.get("item_id"),
                a.get("name"),
                a.get("roi"),
                a.get("sales_per_day"),
                a.get("score"),
                ",".join(a.get("flags", [])),
                a.get("risk"),
            ]
        )
    ws_advice.auto_filter.ref = ws_advice.dimensions
    ws_advice.freeze_panes = "A2"
    _auto_fit(ws_advice)

    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for ws in (ws_items, ws_recipes, ws_advice):
        # openpyxl typing for header/footer is loose; ignore for mypy
        ws.oddFooter.center.text = f"Exported {ts} - World: {world}"  # type: ignore[union-attr]

    return wb
